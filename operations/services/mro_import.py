from pathlib import Path
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.utils import timezone
from masters.models import Planter, SugarMill
from operations.models import MolassesReleaseOrder, normalize_crop_year


def import_mro_workbook(source, default_crop_year="2024 - 25", clear_existing=False):
    """
    Import MRO release orders from Excel workbook (.xlsx/.xls) or CSV file.
    Each sheet in an Excel workbook represents a Sugar Mill / Supplier (e.g. BUSCO, HAWAIIAN, BISCOM, CASA, LOPEZ).
    """
    import openpyxl

    if clear_existing:
        MolassesReleaseOrder.objects.all().delete()

    created_count = 0
    updated_count = 0

    if isinstance(source, (str, bytes, Path)) or hasattr(source, "read"):
        wb = openpyxl.load_workbook(source, data_only=True)
    else:
        wb = source

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))

        # Filter non-empty rows
        valid_rows = [r for r in raw_rows if r and any(c is not None and str(c).strip() != "" for c in r[:8])]
        if not valid_rows:
            continue

        # Clean sheet name for Sugar Mill
        mill_name_clean = sheet_name.strip()
        sugar_mill_obj = _get_or_create_sugar_mill(mill_name_clean)

        # Detect header row
        header_idx = -1
        col_map = {}

        for idx, row in enumerate(valid_rows):
            row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
            if any(k in row_strs for k in ("PLANTERS", "PLANTER", "TONS", "MRO", "MDO", "TRADER")):
                header_idx = idx
                for c_i, h_val in enumerate(row_strs):
                    if "PLANTER" in h_val:
                        col_map["planter"] = c_i
                    elif "TON" in h_val or "VOLUME" in h_val:
                        col_map["tons"] = c_i
                    elif "DATE" in h_val:
                        col_map["date"] = c_i
                    elif "TRADER" in h_val:
                        col_map["trader"] = c_i
                    elif "MRO" in h_val or "MDO" in h_val:
                        col_map["mro"] = c_i
                    elif "CROP" in h_val or "CY" in h_val:
                        col_map["crop_year"] = c_i
                break

        # Fallback column mapping if no header row found
        if header_idx == -1:
            col_map = {"planter": 0, "tons": 1, "date": 2, "trader": 3, "mro": 4, "crop_year": 5}
            data_rows = valid_rows
        else:
            data_rows = valid_rows[header_idx + 1:]

        last_crop_year = default_crop_year

        for row in data_rows:
            if not row or not any(row[:6]):
                continue

            # Extract raw values using col_map
            p_val = _get_cell(row, col_map.get("planter", 0))
            t_val = _get_cell(row, col_map.get("tons", 1))
            d_val = _get_cell(row, col_map.get("date", 2))
            tr_val = _get_cell(row, col_map.get("trader", 3))
            mro_val = _get_cell(row, col_map.get("mro", 4))
            cy_val = _get_cell(row, col_map.get("crop_year"))

            # Skip subtotal / total summary rows or header repeats
            if p_val.upper() in ("PLANTERS", "PLANTER", "TOTAL", "SUBTOTAL") or mro_val.upper() in ("MRO", "MDO"):
                continue

            if not p_val and not mro_val and not t_val:
                continue

            # Parse Tons
            tons_dec = _parse_decimal(t_val)
            if tons_dec is None or tons_dec == Decimal("0"):
                continue

            # Parse MRO number
            mro_num = mro_val.replace(".0", "").strip() if mro_val else ""
            if not mro_num:
                continue

            # Parse Planter Name (Fallback to Mill Name if blank)
            planter_name = p_val if p_val else mill_name_clean
            planter_obj, _ = Planter.objects.get_or_create(
                name=planter_name,
                defaults={"code": planter_name[:20].upper()}
            )

            # Parse Crop Year
            if cy_val and cy_val.strip():
                crop_year = normalize_crop_year(cy_val)
                last_crop_year = crop_year
            else:
                crop_year = normalize_crop_year(last_crop_year or default_crop_year)

            # Parse Release Date
            rel_date = _parse_date(d_val)

            # Trader fallback
            trader_clean = tr_val if tr_val else "HEINDRICH"

            # Create or update MRO record
            mro_obj, created = MolassesReleaseOrder.objects.get_or_create(
                mro_number=mro_num,
                planter=planter_obj,
                tons=tons_dec,
                crop_year=crop_year,
                sugar_mill_name=mill_name_clean,
                defaults={
                    "sugar_mill": sugar_mill_obj,
                    "release_date": rel_date,
                    "trader": trader_clean,
                    "notes": f"Imported from sheet {sheet_name}",
                }
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

    return created_count, updated_count


def _get_cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _parse_decimal(val):
    if not val:
        return None
    cleaned = str(val).replace(",", "").strip()
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    val_str = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None


def _get_or_create_sugar_mill(name):
    name_map = {
        "BUSCO": "BUSCO Sugar Milling Co.",
        "HAWAIIAN": "Hawaiian-Philippine Company",
        "BISCOM": "Binalbagan-Isabela Sugar Co. (BISCOM)",
        "CASA": "Casa Molasses Source",
        "LOPEZ": "Lopez Sugar Central",
    }
    display_name = name_map.get(name.upper(), name)
    mill_obj, _ = SugarMill.objects.get_or_create(
        name=display_name,
        defaults={"location": name}
    )
    return mill_obj
