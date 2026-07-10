from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db import transaction
from django.utils import timezone

from finance.models import FinancialReconciliation, Invoice
from masters.models import Client, LogisticsPartner, SugarMill
from operations.models import LogisticsLedger, PurchaseOrder, TransactionCluster

CLIENT_NAMES = {
    "ADI": "Absolute Distillers Inc.",
    "PROGREEN": "Progreen",
    "GSMI": "Ginebra San Miguel Inc.",
    "BDI": "BDI",
    "EMPERADOR": "Emperador Distillers Inc.",
    "HTC": "Heindrich Trading Corp.",
}

SOURCE_MILL_NAMES = {
    "BUSCO": "BUSCO Sugar Milling Co.",
    "LOPEZ": "Lopez Sugar Mill",
    "CASA": "Casa Molasses Source",
    "CAB": "CAB Source",
    "AABC": "AABC Mill",
    "URC-PASSI": "URC Passi",
    "URC - PASSI": "URC Passi",
    "URC-PASSI": "URC Passi",
    "URC - SONEDCO": "URC SONEDCO",
    "URC-SONEDCO": "URC SONEDCO",
    "URC-SONEDCO ": "URC SONEDCO",
    "HTC TANK": "HTC Tank Storage",
    "HTC": "HTC Internal Tank",
    "SBTI PHILIPA": "SBTI Philippa",
    "ZKYARC": "ZKYARC Source",
    "CASA X CAB": "Casa x CAB Blend",
}


def _clean_str(val):
    if val is None:
        return ""
    return str(val).strip()


def _to_decimal(val):
    if val is None or val == "":
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _parse_si(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    text = str(val).strip()
    if not text or text.upper() in {"ON GOING", "SI"}:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _client_display(code):
    code = _clean_str(code)
    if not code:
        return "Unassigned Customer"
    return CLIENT_NAMES.get(code.upper(), code)


def _mill_display(source):
    source = _clean_str(source)
    if not source:
        return "Unknown Mill"
    key = source.upper()
    for k, v in SOURCE_MILL_NAMES.items():
        if k.upper() == key:
            return v
    return source


def clear_operational_data():
    from audit.models import SystemAuditTrail
    from finance.models import CashVoucher, CapitalLoan, PaymentExpenseMatch

    PaymentExpenseMatch.objects.all().delete()
    FinancialReconciliation.objects.all().delete()
    Invoice.objects.all().delete()
    CashVoucher.objects.all().delete()
    CapitalLoan.objects.all().delete()
    LogisticsLedger.objects.all().delete()
    PurchaseOrder.objects.all().delete()
    TransactionCluster.objects.all().delete()
    Client.objects.all().delete()
    SugarMill.objects.all().delete()
    LogisticsPartner.objects.all().delete()
    SystemAuditTrail.objects.all().delete()


def load_workbook_rows(source):
    import openpyxl

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def parse_excel_to_preview(source):
    rows = load_workbook_rows(source)
    ongoing_section = False
    ongoing_counter = 0
    parsed_rows = []

    summary = {
        'total_rows': 0,
        'duplicates': 0,
        'warnings': 0,
        'total_volume': 0.0,
    }

    notes_buffer = []
    for row in rows:
        if not row:
            continue

        first_cell = _clean_str(row[0]).upper()
        if first_cell == "ON GOING":
            ongoing_section = True
            continue
        if first_cell == "SI" or first_cell == "HEINDRICH TRADING CORPORATION 2026":
            continue

        si = _parse_si(row[0])
        barge = _clean_str(row[2])
        source_mill = _clean_str(row[3])
        customer = _clean_str(row[7])

        # Continuation row (split source line under same invoice block)
        if si is None and not barge and not source_mill:
            continue

        if si is None:
            if not barge and not source_mill:
                continue
            ongoing_counter += 1
            ref = f"ONGOING-{ongoing_counter:03d}"
            status = TransactionCluster.Status.DRAFT
        else:
            ref = f"SI-{si}"
            status = (
                TransactionCluster.Status.DRAFT
                if ongoing_section
                else TransactionCluster.Status.ACTIVE
            )

        purchase_price = _to_decimal(row[4]) or Decimal("0")
        trucking = _to_decimal(row[5]) or Decimal("0")
        barging = _to_decimal(row[6]) or Decimal("0")
        delivered = _to_decimal(row[8])
        received = _to_decimal(row[9])
        selling = _to_decimal(row[11]) or purchase_price
        amount = _to_decimal(row[12]) or Decimal("0")
        inv_date = row[1]

        if isinstance(inv_date, datetime):
            issued_at_str = inv_date.date().isoformat()
            loaded_at_str = inv_date.isoformat()
        else:
            issued_at_str = timezone.now().date().isoformat()
            loaded_at_str = timezone.now().isoformat()

        extra_notes = "; ".join(notes_buffer)
        notes_buffer = []

        volume = delivered or received or Decimal("0")

        row_warnings = []
        is_duplicate = TransactionCluster.objects.filter(reference_code=ref).exists()
        if is_duplicate:
            row_warnings.append(f"Cluster '{ref}' already exists.")
            summary['duplicates'] += 1

        if delivered and received:
            diff = abs(delivered - received)
            if delivered > 0:
                var_pct = (diff / delivered) * 100
                if var_pct > 1.0:
                    row_warnings.append(f"Volume variance ({var_pct:.2f}%) exceeds 1% tolerance.")

        if not customer:
            row_warnings.append("Customer name is empty.")
        if not source_mill:
            row_warnings.append("Source mill is empty.")
        if volume == 0:
            row_warnings.append("Volume (MT) is zero.")

        if row_warnings:
            summary['warnings'] += len(row_warnings)

        parsed_rows.append({
            'reference_code': ref,
            'client_code': customer,
            'client_display': _client_display(customer),
            'mill_code': source_mill,
            'mill_display': _mill_display(source_mill),
            'barge': barge,
            'purchase_price': str(purchase_price),
            'trucking': str(trucking),
            'barging': str(barging),
            'delivered': str(delivered) if delivered else "",
            'received': str(received) if received else "",
            'selling': str(selling),
            'amount': str(amount),
            'issued_at': issued_at_str,
            'loaded_at': loaded_at_str,
            'status': status,
            'extra_notes': extra_notes,
            'is_duplicate': is_duplicate,
            'warnings': row_warnings,
        })
        summary['total_rows'] += 1
        summary['total_volume'] += float(volume)

    summary['ready_to_import'] = summary['total_rows'] - summary['duplicates']
    return {'rows': parsed_rows, 'summary': summary}


def commit_staged_data(parsed_rows):
    imported = 0
    skipped = 0
    with transaction.atomic():
        for r in parsed_rows:
            ref = r['reference_code']
            if TransactionCluster.objects.filter(reference_code=ref).exists():
                skipped += 1
                continue

            purchase_price = Decimal(r['purchase_price'])
            trucking = Decimal(r['trucking'])
            barging = Decimal(r['barging'])
            delivered = Decimal(r['delivered']) if r['delivered'] else None
            received = Decimal(r['received']) if r['received'] else None
            selling = Decimal(r['selling'])
            amount = Decimal(r['amount'])
            issued_at = datetime.fromisoformat(r['issued_at']).date()
            loaded_at = datetime.fromisoformat(r['loaded_at'])
            if timezone.is_naive(loaded_at):
                loaded_at = timezone.make_aware(loaded_at)

            client, _ = Client.objects.get_or_create(name=r['client_display'])
            mill, _ = SugarMill.objects.get_or_create(name=r['mill_display'])
            partner_name = r['barge'] or "TRUCKING"
            partner, _ = LogisticsPartner.objects.get_or_create(
                name=partner_name,
                defaults={"default_freight_rate": barging or trucking},
            )

            cluster = TransactionCluster.objects.create(
                reference_code=ref,
                client=client,
                sugar_mill=mill,
                contract_notes=r['extra_notes'],
                status=r['status'],
            )

            volume = delivered or received or Decimal("0")
            PurchaseOrder.objects.create(
                cluster=cluster,
                volume_mt=volume,
                unit_price=purchase_price,
                terms=f"Selling ₱{selling}" if selling else "",
                approved_at=loaded_at,
            )

            LogisticsLedger.objects.create(
                cluster=cluster,
                partner=partner,
                vessel_id=r['barge'][:100] if r['barge'] else "",
                loaded_volume_mt=delivered or received or Decimal("0"),
                received_volume_mt=received,
                loaded_at=loaded_at,
                received_at=loaded_at if received else None,
                tracking_fees=trucking,
                barge_fees=barging,
            )

            if ref.startswith("SI-"):
                Invoice.objects.create(
                    cluster=cluster,
                    invoice_number=ref,
                    amount=amount,
                    issued_at=issued_at,
                    status=Invoice.Status.ISSUED if amount > 0 else Invoice.Status.DRAFT,
                )

            FinancialReconciliation.objects.create(cluster=cluster)
            imported += 1

    return imported, skipped


def import_htc_summary(path):
    res = parse_excel_to_preview(path)
    return commit_staged_data(res['rows'])
